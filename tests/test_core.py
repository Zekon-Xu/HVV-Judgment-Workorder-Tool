from __future__ import annotations

import json
import os
import tempfile
import unittest
from copy import deepcopy
from io import BytesIO
from pathlib import Path
from unittest import mock

from openpyxl import Workbook
from PIL import Image

from app import default_whitelist as default_whitelist_module
from app import settings_store, template_store, whitelist as whitelist_module
from app.ai_client import AIClient, AIClientError, AIConfig, detect_wire_api, file_to_data_url, normalize_base_url
from app.ai_extract import SYSTEM_EXTRACT, _apply_standard_output, local_extract_text_or_html, smart_extract
from app.batch_engine import jobs_from_paths, jobs_from_text_blob, process_batch
from app.company_networks import CompanyNetworkStore, extract_company_rules_from_file
from app.default_whitelist import company_attribution_lines, company_network_match
from app import project_profiles
from app.extractor import file_to_text, parse_local_file, parse_text
from app.history import HistoryStore
from app.history_sync import (
    HistorySyncError,
    _workbook_urls_from_share_page,
    download_history_workbook,
    normalize_sync_urls,
)
from app.gui import (
    _fit_window_geometry,
    _insert_clipboard_text,
    _parse_indicator_input,
    _record_count_text,
    _restart_environment,
    _whitelist_items_text,
    _work_order_ip_shortcuts,
)
from app.order_builder import WorkOrder, assemble_order
from app.settings_store import (
    DEFAULT_SETTINGS,
    delete_ai_profile,
    normalize_ai_profiles,
    upsert_ai_profile,
)
from app.threatbook import ThreatBookClient, ThreatBookError, indicator_type
from app.whitelist import WhitelistEngine, check_alert_whitelist_gate, prune_redundant_single_ip_rules
from app.whitelist_import import extract_rules_from_file, merge_rules_from_file


class GuiInteractionTests(unittest.TestCase):
    def test_record_count_text_matches_configuration_copy(self) -> None:
        self.assertEqual(_record_count_text(153, note="（仅用于判断内网IP归属）"), "共 153 条记录（仅用于判断内网IP归属）")
        self.assertEqual(_record_count_text(78, 12), "共 78 条记录，当前显示 12 条")

    def test_clipboard_text_replaces_selection_and_keeps_editor_focus(self) -> None:
        class FakeText:
            def __init__(self) -> None:
                self.operations: list[tuple] = []

            def edit_separator(self) -> None:
                self.operations.append(("separator",))

            def tag_ranges(self, tag: str) -> tuple[str, str]:
                self.operations.append(("ranges", tag))
                return ("1.0", "1.4")

            def delete(self, start: str, end: str) -> None:
                self.operations.append(("delete", start, end))

            def insert(self, index: str, text: str) -> None:
                self.operations.append(("insert", index, text))

            def see(self, index: str) -> None:
                self.operations.append(("see", index))

            def focus_set(self) -> None:
                self.operations.append(("focus",))

        target = FakeText()
        _insert_clipboard_text(target, "new")
        self.assertIn(("delete", "sel.first", "sel.last"), target.operations)
        self.assertIn(("insert", "insert", "new"), target.operations)
        self.assertIn(("focus",), target.operations)


class SettingsTests(unittest.TestCase):
    def test_runtime_bootstraps_project_profiles_under_settings(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            bundled_config = root / "bundle" / "settings"
            bundled_profiles = bundled_config / "projects"
            bundled_profiles.mkdir(parents=True)
            (bundled_profiles / "example_project.json").write_text(
                json.dumps({"name": "example_project"}, ensure_ascii=False), encoding="utf-8"
            )
            runtime_config = root / "runtime" / "settings"
            with mock.patch.multiple(
                settings_store,
                CONFIG_DIR=runtime_config,
                SETTINGS_PATH=runtime_config / "settings.json",
                WHITELIST_PATH=runtime_config / "whitelist.json",
                BUNDLE_ROOT=root / "bundle",
                APP_ROOT=root / "runtime",
            ), mock.patch.object(template_store, "TEMPLATES_DIR", runtime_config / "templates"):
                settings_store.ensure_runtime_files()
            self.assertTrue((runtime_config / "projects" / "example_project.json").is_file())

    def test_runtime_migrates_legacy_config_and_project_profiles(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            legacy_config = root / "config"
            legacy_config.mkdir()
            (legacy_config / "settings.json").write_text("{}", encoding="utf-8")
            legacy_profiles = root / "项目配置"
            legacy_profiles.mkdir()
            (legacy_profiles / "legacy.json").write_text(
                json.dumps({"name": "旧项目"}, ensure_ascii=False), encoding="utf-8"
            )
            runtime_config = root / "settings"
            with mock.patch.multiple(
                settings_store,
                APP_ROOT=root,
                CONFIG_DIR=runtime_config,
                SETTINGS_PATH=runtime_config / "settings.json",
                WHITELIST_PATH=runtime_config / "whitelist.json",
                BUNDLE_ROOT=root / "bundle",
            ), mock.patch.object(template_store, "TEMPLATES_DIR", runtime_config / "templates"):
                settings_store.ensure_runtime_files()
            self.assertTrue((runtime_config / "settings.json").is_file())
            self.assertTrue((runtime_config / "projects" / "legacy.json").is_file())
            self.assertFalse(legacy_config.exists())
            self.assertFalse(legacy_profiles.exists())

    def test_runtime_migrates_legacy_history_cache_into_settings(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            legacy = root / "history_cache.json"
            legacy.write_text("[]", encoding="utf-8")
            runtime_config = root / "settings"
            with mock.patch.multiple(
                settings_store,
                APP_ROOT=root,
                CONFIG_DIR=runtime_config,
                SETTINGS_PATH=runtime_config / "settings.json",
                WHITELIST_PATH=runtime_config / "whitelist.json",
                BUNDLE_ROOT=root / "bundle",
            ), mock.patch.object(template_store, "TEMPLATES_DIR", runtime_config / "templates"):
                settings_store.ensure_runtime_files()
            self.assertTrue((runtime_config / "history_cache.json").is_file())
            self.assertFalse(legacy.exists())

    def test_profile_rename_replaces_instead_of_copying(self) -> None:
        data = {
            "ai_profiles": [
                {"name": "One", "base_url": "https://one.example/v1", "api_key": "a", "model": "m1"},
                {"name": "Two", "base_url": "https://two.example/v1", "api_key": "b", "model": "m2"},
            ],
            "ai_active_profile": "One",
        }
        upsert_ai_profile(
            data,
            name="Renamed",
            original_name="One",
            base_url="https://new.example/api/v1",
            api_key="secret",
            model="new-model",
        )
        self.assertEqual([p["name"] for p in data["ai_profiles"]], ["Renamed", "Two"])
        self.assertEqual(data["ai_active_profile"], "Renamed")
        self.assertEqual(data["ai_api_key"], "secret")

    def test_selected_vision_profile_survives_ai_profile_save(self) -> None:
        data = {
            "ai_profiles": [
                {"name": "DeepSeek", "base_url": "https://deepseek.example/v1", "api_key": "", "model": "deepseek"},
                {"name": "Vision", "base_url": "https://vision.example/v1", "api_key": "key", "model": "vision-model"},
            ],
            "ai_active_profile": "DeepSeek",
            "ai_vision_profile": "Vision",
        }
        upsert_ai_profile(
            data,
            name="Vision",
            original_name="Vision",
            base_url="https://vision.example/v1",
            api_key="key",
            model="vision-model-v2",
        )
        self.assertEqual(data["ai_vision_profile"], "Vision")

    def test_profile_names_are_case_insensitively_unique(self) -> None:
        data = {
            "ai_profiles": [
                {"name": "Local", "base_url": "https://a.example/v1", "api_key": "", "model": "a"},
                {"name": "local", "base_url": "https://b.example/v1", "api_key": "", "model": "b"},
            ]
        }
        names = [item["name"] for item in normalize_ai_profiles(data)]
        self.assertEqual(names, ["Local", "local_2"])

    def test_profile_validation_and_delete_guard(self) -> None:
        data: dict = {}
        normalize_ai_profiles(data)
        with self.assertRaises(ValueError):
            upsert_ai_profile(data, name="x", base_url="not-a-url", api_key="", model="m")
        while len(data["ai_profiles"]) > 1:
            self.assertTrue(delete_ai_profile(data, data["ai_profiles"][-1]["name"]))
        self.assertFalse(delete_ai_profile(data, data["ai_profiles"][0]["name"]))

    def test_save_and_load_settings(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "settings.json"
            with mock.patch.object(settings_store, "SETTINGS_PATH", path):
                data = dict(settings_store.DEFAULT_SETTINGS)
                data["theme"] = "light"
                data["ai_profiles"] = [
                    {"name": "Secure", "base_url": "https://api.example/v1", "api_key": "test-secret-value", "model": "m"}
                ]
                data["ai_active_profile"] = "Secure"
                settings_store.save_settings(data)
                loaded = settings_store.load_settings()
            self.assertEqual(loaded["theme"], "light")
            self.assertEqual(loaded["ai_api_key"], "test-secret-value")
            stored = path.read_text(encoding="utf-8")
            self.assertNotIn("test-secret-value", stored)
            self.assertIn("dpapi:", stored)
            json.loads(stored)

    def test_load_removes_legacy_appearance_settings(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "settings.json"
            legacy = dict(settings_store.DEFAULT_SETTINGS)
            legacy.update({"transparency": 0.7, "content_opacity": 0.5, "wallpaper": "old.png"})
            path.write_text(json.dumps(legacy), encoding="utf-8")
            with mock.patch.object(settings_store, "SETTINGS_PATH", path):
                loaded = settings_store.load_settings()
            self.assertNotIn("transparency", loaded)
            self.assertNotIn("content_opacity", loaded)
            self.assertNotIn("wallpaper", loaded)
            stored = json.loads(path.read_text(encoding="utf-8"))
            self.assertNotIn("transparency", stored)
            self.assertNotIn("content_opacity", stored)
            self.assertNotIn("wallpaper", stored)

    def test_default_window_is_centered_and_never_fullscreen(self) -> None:
        self.assertEqual(_fit_window_geometry("1600x1000+0+0", 1920, 1080), "1600x984+160+36")
        self.assertEqual(_fit_window_geometry("1600x1000", 2560, 1440), "1600x1000+480+208")


class ApiClientTests(unittest.TestCase):
    def test_api_type_auto_detection(self) -> None:
        self.assertEqual(detect_wire_api("https://api.anthropic.com/v1", "claude-sonnet"), "anthropic")
        self.assertEqual(detect_wire_api("https://gateway.example/v1", "gpt-5.6-luna"), "responses")
        self.assertEqual(detect_wire_api("https://api.moonshot.ai/v1", "kimi-k2.6"), "chat")

    def test_responses_and_anthropic_payload_adapters(self) -> None:
        messages = [{"role": "system", "content": "judge"}, {"role": "user", "content": "hello"}]
        response_input = AIClient._responses_input(messages)
        self.assertEqual(response_input[1]["content"][0]["type"], "input_text")
        anthropic = AIClient._anthropic_payload(messages, "claude-test", 200, 0.1)
        self.assertEqual(anthropic["system"], "judge")
        self.assertEqual(anthropic["messages"][0]["role"], "user")

    def test_base_url_normalization(self) -> None:
        self.assertEqual(normalize_base_url("https://api.example.com"), "https://api.example.com/v1")
        self.assertEqual(
            normalize_base_url("https://api.example.com/api/v1/chat/completions?x=1"),
            "https://api.example.com/api/v1",
        )
        self.assertEqual(
            normalize_base_url("http://127.0.0.1:11434/v1/"),
            "http://127.0.0.1:11434/v1",
        )

    def test_default_prompt_contains_required_sample(self) -> None:
        self.assertIn("编号：0807-165", SYSTEM_EXTRACT)
        self.assertIn("攻击IP：103.213.96.237", SYSTEM_EXTRACT)
        self.assertIn("没有证据的值留空", SYSTEM_EXTRACT)
        self.assertIn("当前模板工单初版", SYSTEM_EXTRACT)

    def test_png_data_url_preserves_png(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "alpha.png"
            Image.new("RGBA", (4, 4), (255, 0, 0, 100)).save(path)
            self.assertTrue(file_to_data_url(path).startswith("data:image/png;base64,"))

    def test_deepseek_v4_requests_disable_thinking(self) -> None:
        response = mock.Mock(status_code=200)
        response.json.return_value = {"choices": [{"message": {"content": "ok"}, "finish_reason": "stop"}]}
        with mock.patch("app.ai_client.httpx.Client") as client_type:
            client_type.return_value.__enter__.return_value.post.return_value = response
            client = AIClient(AIConfig(
                api_key="key", base_url="https://api.deepseek.com/v1",
                model="deepseek-v4-flash", wire_api="chat",
            ))
            self.assertEqual(client.chat([{"role": "user", "content": "ok"}]), "ok")
            payload = client_type.return_value.__enter__.return_value.post.call_args.kwargs["json"]
        self.assertEqual(payload["thinking"], {"type": "disabled"})


class ThreatBookTests(unittest.TestCase):
    def test_manual_indicator_input_is_split_and_deduplicated(self) -> None:
        self.assertEqual(
            _parse_indicator_input("198.51.100.8, 203.0.113.4\n198.51.100.8"),
            ["198.51.100.8", "203.0.113.4"],
        )

    def test_work_order_shortcuts_include_each_ip_from_standard_and_custom_fields(self) -> None:
        order = WorkOrder(
            attack_ip="198.51.100.8",
            target_ip="10.2.3.4",
            xff="198.51.100.8, 203.0.113.4",
            custom_fields={"证据": "备用地址 192.0.2.9"},
        )
        self.assertEqual(
            _work_order_ip_shortcuts(order, "附加 IP 203.0.113.4"),
            ["198.51.100.8", "10.2.3.4", "203.0.113.4", "192.0.2.9"],
        )

    def test_indicator_validation(self) -> None:
        self.assertEqual(indicator_type("198.51.100.8"), "ip")
        self.assertEqual(indicator_type("example.org"), "domain")
        with self.assertRaises(ThreatBookError):
            indicator_type("not an indicator")

    def test_ip_lookup_uses_reputation_endpoint(self) -> None:
        response = mock.Mock(status_code=200)
        response.json.return_value = {"response_code": 0, "data": {"judgment": "malicious", "tags": ["scanner"]}}
        settings = {"threatbook_enabled": True, "threatbook_api_key": "key", "threatbook_timeout": 5}
        with mock.patch("app.threatbook.httpx.get", return_value=response) as get:
            result = ThreatBookClient(settings).lookup("198.51.100.8")
        self.assertIn("malicious", result.summary)
        self.assertIn("scanner", result.summary)
        self.assertIn("scene/ip_reputation", get.call_args.args[0])

    def test_domain_lookup_uses_domain_query_endpoint(self) -> None:
        response = mock.Mock(status_code=200)
        response.json.return_value = {"response_code": 0, "data": {"zdg16881988.com": {"judgments": ["Malware"]}}}
        settings = {"threatbook_enabled": True, "threatbook_api_key": "key"}
        with mock.patch("app.threatbook.httpx.get", return_value=response) as get:
            result = ThreatBookClient(settings).lookup("zdg16881988.com")
        self.assertIn("domain/query", get.call_args.args[0])
        self.assertIn("Malware", result.summary)
        self.assertIn("https://x.threatbook.com/v5/domain/zdg16881988.com", result.summary)

    def test_single_ip_lookup_matches_containing_cidr_and_prunes_redundant_rule(self) -> None:
        rules, removed = prune_redundant_single_ip_rules([
            {"rule": "124.238.251.139", "reason": "单条"},
            {"rule": "124.238.251.128/25", "reason": "云防出口IP"},
        ])
        self.assertEqual(removed, 1)
        self.assertEqual([item["rule"] for item in rules], ["124.238.251.128/25"])
        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "whitelist.json"
            path.write_text(json.dumps({"rules": [{"rule": "124.238.251.128/25", "reason": "云防出口IP"}]}), encoding="utf-8")
            engine = WhitelistEngine(path)
            match = engine.check("124.238.251.139")
            self.assertTrue(match.matched)
            self.assertEqual(match.rule, "124.238.251.128/25")


class TemplateTests(unittest.TestCase):
    def _patch_roots(self, root: Path):
        return mock.patch.multiple(
            template_store,
            TEMPLATES_DIR=root / "config" / "templates",
            APP_ROOT=root,
            BUNDLE_ROOT=root,
        )

    def test_first_run_creates_default_without_recursion(self) -> None:
        with tempfile.TemporaryDirectory() as tmp, self._patch_roots(Path(tmp)):
            directory = template_store.ensure_templates_dir()
            self.assertTrue((directory / "default.json").is_file())
            self.assertEqual(template_store.list_templates()[0]["name"], template_store.BUILTIN_TEMPLATE_NAME)

    def test_delete_user_template_but_not_builtin(self) -> None:
        with tempfile.TemporaryDirectory() as tmp, self._patch_roots(Path(tmp)):
            template_store.ensure_templates_dir()
            template_store.save_template({"name": "临时模板", "field_schema": {"attack_ip": "IP"}})
            self.assertTrue(template_store.delete_template("临时模板"))
            with self.assertRaises(ValueError):
                template_store.delete_template(template_store.BUILTIN_TEMPLATE_NAME)

    def test_manual_template_field_can_add_remove_and_reorder(self) -> None:
        with tempfile.TemporaryDirectory() as tmp, self._patch_roots(Path(tmp)):
            template_store.ensure_templates_dir()
            template_store.add_manual_template_field(template_store.BUILTIN_TEMPLATE_NAME, "字段A", "值A")
            template_store.add_manual_template_field(template_store.BUILTIN_TEMPLATE_NAME, "攻击IP", "203.0.113.8")
            template_store.add_manual_template_field(template_store.BUILTIN_TEMPLATE_NAME, "字段C", "值C")
            _path, moved = template_store.move_template_field(
                template_store.BUILTIN_TEMPLATE_NAME, "字段C", 0
            )
            self.assertEqual(list(moved["field_schema"]), ["字段C", "字段A", "攻击IP"])
            self.assertEqual(moved["field_bindings"]["攻击IP"], "attack_ip")
            _path, removed = template_store.remove_template_field(
                template_store.BUILTIN_TEMPLATE_NAME, "字段A"
            )
            self.assertEqual(list(removed["field_schema"]), ["字段C", "攻击IP"])
            self.assertNotIn("字段A", removed["sample_fields"])

    def test_manual_field_persists_rows_and_custom_options(self) -> None:
        with tempfile.TemporaryDirectory() as tmp, self._patch_roots(Path(tmp)):
            template_store.ensure_templates_dir()
            _path, added = template_store.add_manual_template_field(
                template_store.BUILTIN_TEMPLATE_NAME,
                "审批状态",
                "待确认",
                rows=4,
                options=["待确认", "已确认", "待确认"],
            )
            self.assertEqual(added["field_rows"]["审批状态"], 4)
            self.assertEqual(added["field_options"]["审批状态"], ["待确认", "已确认"])
            loaded = template_store.load_template(template_store.BUILTIN_TEMPLATE_NAME)
            self.assertEqual(loaded["field_rows"]["审批状态"], 4)
            self.assertEqual(loaded["field_options"]["审批状态"], ["待确认", "已确认"])
            _path, moved = template_store.move_template_field(
                template_store.BUILTIN_TEMPLATE_NAME, "审批状态", 0
            )
            self.assertEqual(moved["field_rows"]["审批状态"], 4)
            _path, removed = template_store.remove_template_field(
                template_store.BUILTIN_TEMPLATE_NAME, "审批状态"
            )
            self.assertNotIn("审批状态", removed["field_rows"])
            self.assertNotIn("审批状态", removed["field_options"])

    def test_import_schema_is_exact_and_does_not_overwrite(self) -> None:
        with tempfile.TemporaryDirectory() as tmp, self._patch_roots(Path(tmp)):
            root = Path(tmp)
            template_store.ensure_templates_dir()
            source = root / "incoming.json"
            source.write_text(
                json.dumps({"name": "导入模板", "field_schema": {"attack_ip": "custom"}}, ensure_ascii=False),
                encoding="utf-8",
            )
            first = template_store.import_template_file(source)
            second = template_store.import_template_file(source)
            self.assertNotEqual(first, second)
            loaded = template_store.load_template(first)
            self.assertEqual(loaded["field_schema"]["attack_ip"], "custom")
            self.assertEqual(loaded["field_schema"], {"attack_ip": "custom"})
            _path, expanded = template_store.add_manual_template_field(first, "攻击次数", "3")
            self.assertEqual(list(expanded["field_schema"]), ["attack_ip", "攻击次数"])
            _path, reordered = template_store.move_template_field(first, "攻击次数", 0)
            self.assertEqual(list(reordered["field_schema"]), ["攻击次数", "attack_ip"])
            _path, reduced = template_store.remove_template_field(first, "attack_ip")
            self.assertEqual(list(reduced["field_schema"]), ["攻击次数"])

    def test_sample_template_preserves_only_supplied_fields_and_order(self) -> None:
        template = template_store.template_from_sample(
            "自定义四字段",
            "编1号：0810-010\n监2：360一级\n3：2026-08-10 01:36:55\n4：182.126.123.181",
        )
        self.assertEqual(list(template["field_schema"]), ["编1号", "监2", "3", "4"])
        self.assertEqual(template["sample_fields"]["4"], "182.126.123.181")
        self.assertEqual(template["field_bindings"]["编1号"], "number")
        self.assertEqual(template["field_bindings"]["监2"], "source")
        self.assertEqual(template["field_bindings"]["3"], "time")
        self.assertNotIn("4", template["field_bindings"])
        self.assertEqual(len(template["local_rules"]), 4)
        self.assertIn("编1号、监2、3、4", template["ai_prompt"])

    def test_generated_local_rules_parse_nonstandard_template_labels(self) -> None:
        sample = "编1号：0810-010\n监2：360一级\n3：2026-08-10 01:36:55\n4：182.126.123.181"
        template = template_store.template_from_sample("非标准四字段", sample)
        alert = local_extract_text_or_html(text=sample, template=template)
        self.assertEqual(alert.template_fields, {
            "编1号": "0810-010", "监2": "360一级",
            "3": "2026-08-10 01:36:55", "4": "182.126.123.181",
        })
        self.assertEqual(alert.time, "2026-08-10 01:36:55")

    def test_template_rules_have_no_legacy_field_count_limit(self) -> None:
        sample = "\n".join(f"字段{index}：值{index}" for index in range(1, 41))
        template = template_store.template_from_sample("四十字段", sample)
        alert = local_extract_text_or_html(text=sample, template=template)
        self.assertEqual(len(template["local_rules"]), 40)
        self.assertEqual(len(alert.template_fields), 40)
        self.assertEqual(alert.template_fields["字段40"], "值40")

    def test_standard_labels_receive_semantic_bindings(self) -> None:
        template = template_store.template_from_sample(
            "标准字段", "攻击IP：1.2.3.4\n目标IP：10.2.3.4\n处置建议：待研判"
        )
        self.assertEqual(template["field_bindings"]["攻击IP"], "attack_ip")
        self.assertEqual(template["field_bindings"]["目标IP"], "target_ip")

    def test_ai_output_keeps_template_draft_and_standard_fields(self) -> None:
        template = {
            "field_schema": {"自定义源地址": "自定义源地址", "攻击次数": "攻击次数"}
        }
        output = """【当前模板工单初版】
自定义源地址：203.0.113.8
攻击次数：12
【标准研判字段】
攻击IP：203.0.113.8
目标IP：10.2.3.4
攻击名称：SQL注入攻击
事件类型：SQL注入
攻击结果：失败
研判依据：请求中存在明确注入特征
处置依据：攻击源非白名单"""
        alert = _apply_standard_output(output, template=template)
        self.assertEqual(alert.attack_ip, "203.0.113.8")
        self.assertEqual(alert.template_fields["攻击次数"], "12")
        self.assertTrue(any("AI研判依据" in note for note in alert.notes))

    def test_template_attack_attempt_is_normalized_to_failure(self) -> None:
        template = {
            "field_schema": {"攻击结果": "攻击结果"},
            "field_bindings": {"攻击结果": "attack_result"},
        }
        alert = _apply_standard_output(
            "【当前模板工单初版】\n攻击结果：企图\n【标准研判字段】\n攻击IP：203.0.113.8\n攻击名称：扫描攻击\n攻击结果：企图",
            template=template,
        )
        self.assertEqual(alert.attack_result, "失败")
        self.assertEqual(alert.template_fields["攻击结果"], "失败")


class ProjectProfileTests(unittest.TestCase):
    def _patch_paths(self, root: Path):
        return mock.patch.multiple(
            project_profiles,
            PROJECT_PROFILES_DIR=root / "项目配置",
            SETTINGS_PATH=root / "config" / "settings.json",
            WHITELIST_PATH=root / "config" / "whitelist.json",
            HISTORY_CACHE_PATH=root / "config" / "history_cache.json",
            TEMPLATES_DIR=root / "config" / "templates",
        )

    def test_save_and_restore_project_bundle(self) -> None:
        with tempfile.TemporaryDirectory() as tmp, self._patch_paths(Path(tmp)):
            root = Path(tmp)
            with mock.patch.object(settings_store, "SETTINGS_PATH", root / "config" / "settings.json"), \
                 mock.patch.object(template_store, "TEMPLATES_DIR", root / "config" / "templates"), \
                 mock.patch.object(default_whitelist_module, "SETTINGS_PATH", root / "config" / "settings.json"), \
                 mock.patch.object(default_whitelist_module, "CONFIG_DIR", root / "config"):
                (root / "config" / "templates").mkdir(parents=True)
                (root / "config" / "templates" / "default.json").write_text(
                    json.dumps({"name": "项目模板", "field_schema": {"字段A": "字段A"}}, ensure_ascii=False),
                    encoding="utf-8",
                )
                (root / "config" / "whitelist.json").write_text(
                    json.dumps({"version": 1, "rules": [{"rule": "192.0.2.1", "reason": "test"}], "manual": []}),
                    encoding="utf-8",
                )
                settings = deepcopy(DEFAULT_SETTINGS)
                settings["app_title"] = "项目A"
                profile_path = project_profiles.save_project_profile("2027项目AHVV", settings)
                profile = json.loads(profile_path.read_text(encoding="utf-8"))
                self.assertEqual(profile["company_networks"]["rules"], [])
                project_profiles.restore_blank_workspace()
                loaded = project_profiles.load_project_profile(profile_path)
                restored = settings_store.load_settings()
            self.assertEqual(loaded["name"], "2027项目AHVV")
            self.assertEqual(restored["app_title"], "项目A")
            self.assertEqual(restored["active_project_profile"], "2027项目AHVV")
            whitelist = json.loads((root / "config" / "whitelist.json").read_text(encoding="utf-8"))
            self.assertEqual(whitelist["rules"][0]["rule"], "192.0.2.1")

    def test_project_bundle_sensitive_options_are_portable_and_selective(self) -> None:
        with tempfile.TemporaryDirectory() as tmp, self._patch_paths(Path(tmp)):
            root = Path(tmp)
            (root / "config" / "templates").mkdir(parents=True)
            (root / "config" / "templates" / "default.json").write_text(
                json.dumps({"name": "默认", "field_schema": {}}, ensure_ascii=False), encoding="utf-8"
            )
            (root / "config" / "whitelist.json").write_text(
                json.dumps({"version": 1, "rules": [{"rule": "192.0.2.1"}]}), encoding="utf-8"
            )
            (root / "config" / "history_cache.json").write_text(
                json.dumps([{"attack_ip": "198.51.100.8"}]), encoding="utf-8"
            )
            settings = deepcopy(DEFAULT_SETTINGS)
            settings["ai_api_key"] = "portable-ai"
            settings["ai_profiles"][0]["api_key"] = "portable-ai"
            settings["threatbook_api_key"] = "portable-ti"
            path = project_profiles.save_project_profile(
                "portable", settings,
                include_ai_key=True,
                include_threatbook_key=False,
                include_whitelist=False,
                include_company_networks=False,
            )
            data = json.loads(path.read_text(encoding="utf-8"))
            self.assertEqual(data["settings"]["ai_api_key"], "portable-ai")
            self.assertEqual(data["settings"]["threatbook_api_key"], "")
            self.assertEqual(data["whitelist"]["rules"], [])
            self.assertEqual(data["company_networks"]["rules"], [])
            self.assertEqual(data["history_cache"][0]["attack_ip"], "198.51.100.8")
            self.assertFalse(str(data["settings"]["ai_api_key"]).startswith("dpapi:"))

    def test_frozen_restart_uses_a_fresh_pyinstaller_temp_directory(self) -> None:
        with mock.patch.dict(os.environ, {
            "_PYI_APPLICATION_HOME_DIR": r"C:\Temp\_MEI123",
            "_MEIPASS2": r"C:\Temp\_MEI123",
        }, clear=False):
            env = _restart_environment()
        self.assertEqual(env["PYINSTALLER_RESET_ENVIRONMENT"], "1")
        self.assertNotIn("_PYI_APPLICATION_HOME_DIR", env)
        self.assertNotIn("_MEIPASS2", env)


class WhitelistTests(unittest.TestCase):
    def setUp(self) -> None:
        self._settings_tmp = tempfile.TemporaryDirectory()
        settings_path = Path(self._settings_tmp.name) / "settings.json"
        settings_path.write_text(
            json.dumps({"company_networks_blank": True, "active_project_profile": ""}),
            encoding="utf-8",
        )
        self._settings_patch = mock.patch.object(default_whitelist_module, "SETTINGS_PATH", settings_path)
        self._config_patch = mock.patch.object(default_whitelist_module, "CONFIG_DIR", Path(self._settings_tmp.name))
        self._settings_patch.start()
        self._config_patch.start()

    def tearDown(self) -> None:
        self._config_patch.stop()
        self._settings_patch.stop()
        self._settings_tmp.cleanup()

    def _engine(self, root: Path, rules: list[dict] | None = None) -> WhitelistEngine:
        path = root / "whitelist.json"
        path.write_text(json.dumps({"version": 1, "rules": rules or [], "manual": []}, ensure_ascii=False), encoding="utf-8")
        return WhitelistEngine(path)

    def test_company_networks_only_attribute_departments(self) -> None:
        self.assertIsNone(company_network_match("10.2.18.9"))
        self.assertEqual(company_attribution_lines(attack_ip="10.2.18.9"), [])

    def test_role_aware_whitelist_gate(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            engine = self._engine(Path(tmp), [
                {"rule": "192.0.2.0/24", "reason": "出口"},
                {"rule": "198.51.100.8", "reason": "代理"},
            ])
            self.assertTrue(check_alert_whitelist_gate(engine, "192.0.2.9").skip_order)
            self.assertTrue(check_alert_whitelist_gate(engine, "192.0.2.9", xff="198.51.100.8").skip_order)
            self.assertFalse(check_alert_whitelist_gate(engine, "192.0.2.9", xff="203.0.113.8").skip_order)
            self.assertFalse(check_alert_whitelist_gate(engine, "203.0.113.8").skip_order)
            self.assertFalse(check_alert_whitelist_gate(engine, "", xff="198.51.100.8").skip_order)
            self.assertFalse(check_alert_whitelist_gate(engine, "192.0.2.9", target_ip="10.3.2.2").skip_order)

    def test_domain_ioc_requires_explicit_whitelist(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            engine = self._engine(Path(tmp), [
                {"rule": "203.0.113.0/24", "reason": "出口"},
                {"rule": "safe.example.com", "reason": "可信域名"},
            ])
            self.assertFalse(check_alert_whitelist_gate(engine, "203.0.113.8", domain_url="zdg16881988.com").skip_order)
            self.assertTrue(check_alert_whitelist_gate(engine, "203.0.113.8", domain_url="safe.example.com").skip_order)

    def test_company_attack_ip_is_only_a_semi_whitelist(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            engine = self._engine(Path(tmp), [{"rule": "198.51.100.8", "reason": "目标白名单"}])
            allowed = check_alert_whitelist_gate(
                engine, attack_ip="10.2.18.9", target_ip="198.51.100.8"
            )
            target_not_allowed = check_alert_whitelist_gate(
                engine, attack_ip="10.2.18.9", target_ip="10.3.2.2"
            )
            self.assertFalse(allowed.skip_order)
            self.assertEqual(allowed.semi_matched, [])
            self.assertFalse(target_not_allowed.skip_order)
        self.assertEqual(target_not_allowed.unmatched[0]["role"], "攻击IP")

    def test_whitelist_dialog_uses_role_specific_brackets(self) -> None:
        text = _whitelist_items_text([
            {"role": "攻击IP", "ip": "198.51.100.1", "reason": "出口"},
            {"role": "目标/受害/目的IP", "ip": "10.2.3.4", "reason": "目的"},
        ], reasons=True)
        self.assertIn("【攻击IP 198.51.100.1】", text)
        self.assertIn("【目的IP 10.2.3.4】", text)

    def test_manual_rules_validate_deduplicate_and_delete(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            engine = self._engine(Path(tmp))
            self.assertTrue(engine.add_manual("192.0.2.10/24", "test"))
            self.assertFalse(engine.add_manual("192.0.2.0/24", "test"))
            self.assertTrue(engine.check("192.0.2.99").matched)
            with self.assertRaises(ValueError):
                engine.add_manual("10.2.3.0-4.0", "bad")
            self.assertTrue(engine.remove_manual("192.0.2.0/24"))
            self.assertFalse(engine.remove_manual("192.0.2.0/24"))

    def test_all_rules_are_editable_and_json_backup_can_restore(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            path = root / "whitelist.json"
            path.write_text(json.dumps({
                "version": 1,
                "rules": [{"rule": "192.0.2.1", "source": "导入"}],
                "manual": [{"rule": "198.51.100.1", "source": "旧手动"}],
            }), encoding="utf-8")
            engine = WhitelistEngine(path)
            self.assertEqual(len(engine.all_entries()), 2)
            self.assertTrue(engine.remove_rule("192.0.2.1"))
            backup = root / "backup.json"
            engine.export_json(backup)
            self.assertTrue(engine.remove_rule("198.51.100.1"))
            self.assertEqual(engine.restore_json(backup), 1)
            self.assertTrue(engine.check("198.51.100.1").matched)
            saved = json.loads(path.read_text(encoding="utf-8"))
            self.assertNotIn("manual", saved)

    def test_file_import_is_incremental_and_deduplicated(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            engine = self._engine(root)
            source = root / "whitelist.txt"
            source.write_text("出口A 192.0.2.1\n出口B 198.51.100.0/24\n重复 192.0.2.1", encoding="utf-8")
            self.assertEqual(merge_rules_from_file(engine, source), 2)
            self.assertEqual(merge_rules_from_file(engine, source), 0)
            self.assertTrue(engine.check("198.51.100.8").matched)

    def test_xlsx_import_never_promotes_company_network_sheet_to_whitelist(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "公司网段及出口信息表.xlsx"
            workbook = Workbook()
            company = workbook.active
            company.title = "公司网段"
            company.append(["公司名称", "简称", "网段"])
            company.append(["示例公司", "示例部门", "10.2.0.0"])
            whitelist = workbook.create_sheet("出口白名单")
            whitelist.append(["厦门出口", "北京出口", "DNS", "VPN", "防火墙", "扫描器", "渗透固定ip"])
            whitelist.append(["192.0.2.0/30", "198.51.100.0/26", "10.1.1.5", "10.255.0.0/16", "10.201.130.2", "10.10.207.33", "示例账号"])
            workbook.save(path)

            rules = extract_rules_from_file(path)
            imported = {item["rule"] for item in rules}
            self.assertIn("192.0.2.0/30", imported)
            self.assertIn("10.1.1.5", imported)
            self.assertNotIn("10.2.0.0", imported)

            company_only = Path(tmp) / "仅公司网段.xlsx"
            workbook = Workbook()
            workbook.active.append(["公司名称", "简称", "网段"])
            workbook.active.append(["示例公司", "示例部门", "10.2.0.0"])
            workbook.save(company_only)
            with self.assertRaises(ValueError):
                extract_rules_from_file(company_only)

    def test_company_network_import_expands_ranges_and_store_can_rollback(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            settings_path = root / "settings.json"
            projects = root / "projects"
            projects.mkdir()
            settings_path.write_text(
                json.dumps({"active_project_profile": "测试项目"}, ensure_ascii=False), encoding="utf-8"
            )
            profile_path = projects / "测试项目.json"
            profile_path.write_text(json.dumps({
                "name": "测试项目",
                "settings": {},
                "company_networks": {"rules": [
                    {"rule": "10.2.0.0/16", "reason": "示例部门", "source": "测试"}
                ]},
            }, ensure_ascii=False), encoding="utf-8")
            source = root / "公司网段.xlsx"
            workbook = Workbook()
            workbook.active.append(["公司名称", "简称", "网段"])
            workbook.active.append(["云平台", "云平台", "10.210.11.0-13.0"])
            workbook.save(source)

            imported = extract_company_rules_from_file(source)
            self.assertEqual(
                {item["rule"] for item in imported},
                {"10.210.11.0/24", "10.210.12.0/24", "10.210.13.0/24"},
            )
            with mock.patch.object(default_whitelist_module, "SETTINGS_PATH", settings_path), \
                 mock.patch.object(default_whitelist_module, "CONFIG_DIR", root):
                store = CompanyNetworkStore()
                self.assertEqual(store.merge_rules(imported), 3)
                backup = root / "company-backup.json"
                store.export_json(backup)
                self.assertTrue(store.remove("10.210.12.0/24"))
                self.assertEqual(store.restore_json(backup), 4)
                self.assertEqual(len(store.all_entries()), 4)

    def test_restore_original_replaces_custom_rules(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            baseline = root / "original.json"
            baseline.write_text(json.dumps({
                "description": "test baseline",
                "rules": [{"rule": "198.51.100.0/24", "reason": "source", "source": "测试"}],
            }), encoding="utf-8")
            engine = self._engine(root)
            engine.add_manual("192.0.2.1")
            with mock.patch.object(whitelist_module, "ORIGINAL_WHITELIST_PATH", baseline):
                self.assertEqual(engine.restore_original(), 1)
            self.assertTrue(engine.check("198.51.100.9").matched)
            self.assertFalse(engine.check("192.0.2.1").matched)


class HistorySyncTests(unittest.TestCase):
    def test_confirmed_order_is_incremental(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            store = HistoryStore(Path(tmp) / "history.json")
            order = WorkOrder(number="0810-001", attack_ip="198.51.100.8", target_ip="10.2.3.4")
            changed, total = store.confirm_order(order)
            self.assertTrue(changed)
            self.assertEqual(total, 1)
            changed, total = store.confirm_order(order)
            self.assertFalse(changed)
            self.assertEqual(total, 1)
            self.assertEqual(store.find_exact_code("0810-001")["all_fields"]["number"], "0810-001")
            self.assertEqual(len(store.search("198.51.100.8")), 1)

    def test_remote_xlsx_merge_updates_same_number(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            store = HistoryStore(Path(tmp) / "history.json")
            workbook = Workbook()
            sheet = workbook.active
            sheet.append(["序号", "编号", "监测来源", "时间", "攻击IP", "目标IP", "XFF", "域名URL", "攻击名称", "事件类型", "上报人员"])
            sheet.append([1, "0810-001", "源A", "2026-08-10", "198.51.100.8", "10.2.3.4", "", "", "扫描", "工具扫描", "A"])
            payload = BytesIO()
            workbook.save(payload)
            self.assertEqual(store.merge_from_xlsx_bytes(payload.getvalue())[:2], (1, 0))
            workbook = Workbook()
            sheet = workbook.active
            sheet.append(["序号", "编号", "监测来源", "时间", "攻击IP", "目标IP", "XFF", "域名URL", "攻击名称", "事件类型", "上报人员"])
            sheet.append([1, "0810-001", "源B", "2026-08-10", "198.51.100.8", "10.2.3.4", "", "", "扫描", "工具扫描", "B"])
            payload = BytesIO()
            workbook.save(payload)
            self.assertEqual(store.merge_from_xlsx_bytes(payload.getvalue())[:2], (0, 1))
            self.assertEqual(store.records[0]["source"], "源B")
            self.assertIn("攻击IP", store.records[0]["all_fields"])

    def test_url_normalization_and_login_page_rejection(self) -> None:
        self.assertEqual(normalize_sync_urls("https://a.example/a.xlsx\ninvalid\nhttps://a.example/a.xlsx"), ["https://a.example/a.xlsx"])

        class Response:
            url = "https://account.kdocs.cn/passport/singlesign"
            headers = {"content-type": "text/html"}
            content = b"<html>login</html>"
            def raise_for_status(self):
                return None

        with mock.patch("app.history_sync.httpx.get", return_value=Response()):
            with self.assertRaises(HistorySyncError):
                download_history_workbook("https://www.kdocs.cn/l/example")

    def test_share_page_only_uses_trusted_explicit_excel_exports(self) -> None:
        page = (
            '"https:\\/\\/docs.kdocs.cn\\/download\\/tracking.xlsx?token=abc" '
            '"https:\\/\\/example.invalid\\/download\\/ignored.xlsx"'
        )
        self.assertEqual(
            _workbook_urls_from_share_page(page),
            ["https://docs.kdocs.cn/download/tracking.xlsx?token=abc"],
        )


class WorkflowTests(unittest.TestCase):
    SAMPLE = """时间：2026-08-07 21:58:00
源IP：8.8.8.8
目的IP：10.2.3.4
攻击名称：SSH暴力破解攻击
告警级别：高危
攻击结果：失败
"""

    def _empty_engine(self, root: Path) -> WhitelistEngine:
        path = root / "wl.json"
        path.write_text('{"rules": [], "manual": []}', encoding="utf-8")
        return WhitelistEngine(path)

    def test_local_extraction_roles_and_ai_switch(self) -> None:
        alert = parse_text(self.SAMPLE)
        self.assertEqual(alert.attack_ip, "8.8.8.8")
        self.assertEqual(alert.target_ip, "10.2.3.4")
        result = smart_extract(
            settings={"ai_enabled": True, "ai_use_ocr": False, "ai_api_key": "key", "ai_base_url": "https://a.example/v1", "ai_model": "m"},
            text=self.SAMPLE,
        )
        self.assertEqual(result.attack_ip, "8.8.8.8")
        self.assertTrue(any("已关闭" in note for note in result.notes))

    def test_force_local_never_calls_ai(self) -> None:
        settings = {
            "analysis_mode": "local",
            "ai_enabled": True,
            "ai_use_ocr": True,
            "ai_api_key": "configured",
            "ai_base_url": "https://api.example/v1",
            "ai_model": "model",
        }
        with mock.patch("app.ai_extract.AIClient") as client:
            alert = smart_extract(settings=settings, text=self.SAMPLE)
        client.assert_not_called()
        self.assertEqual(alert.attack_ip, "8.8.8.8")
        self.assertTrue(any("本地文本规则解析" in note for note in alert.notes))

    def test_force_ai_does_not_silently_fallback(self) -> None:
        with self.assertRaises(AIClientError):
            smart_extract(
                settings={"analysis_mode": "ai", "ai_enabled": True},
                text=self.SAMPLE,
            )

    def test_force_local_rejects_images_without_ocr(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "alert.png"
            Image.new("RGB", (20, 20), "white").save(path)
            with self.assertRaisesRegex(AIClientError, "图片必须切换"):
                smart_extract(
                    settings={"analysis_mode": "local", "ai_enabled": False},
                    path=path,
                )

    def test_local_html_table_and_custom_aliases(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "alert.html"
            path.write_text(
                """<table>
                <tr><th>客户端地址</th><td>8.8.4.4</td></tr>
                <tr><th>受影响地址</th><td>10.2.9.8</td></tr>
                <tr><th>告警名称</th><td>SQL注入攻击</td></tr>
                <tr><th>攻击结果</th><td>已拦截</td></tr>
                </table>""",
                encoding="utf-8",
            )
            template = {
                "field_aliases": {
                    "attack_ip": ["客户端地址"],
                    "target_ip": ["受影响地址"],
                }
            }
            alert = local_extract_text_or_html(path=path, template=template)
        self.assertEqual(alert.attack_ip, "8.8.4.4")
        self.assertEqual(alert.target_ip, "10.2.9.8")
        self.assertEqual(alert.attack_name, "SQL注入攻击")
        self.assertEqual(alert.event_type, "SQL注入")
        self.assertEqual(alert.attack_result, "失败")

    def test_local_json_and_csv_are_flattened(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            json_path = root / "alert.json"
            json_path.write_text(
                json.dumps(
                    {
                        "event": {
                            "timestamp": "2026-08-09T03:04:05",
                            "sensor_ip": "192.0.2.40",
                            "src_ip": "1.2.3.4",
                            "dst_ip": "10.3.2.1",
                            "rule_name": "Nmap scan",
                            "severity": "medium",
                        }
                    }
                ),
                encoding="utf-8",
            )
            csv_path = root / "alert.csv"
            csv_path.write_text(
                "src_ip,dst_ip,alert_name,result\n9.9.9.9,10.4.5.6,SSH brute force,blocked\n",
                encoding="utf-8",
            )
            json_alert = parse_local_file(json_path)
            csv_alert = parse_local_file(csv_path)
        self.assertEqual(json_alert.attack_ip, "1.2.3.4")
        self.assertEqual(json_alert.target_ip, "10.3.2.1")
        self.assertEqual(json_alert.alert_level, "中危")
        self.assertEqual(json_alert.event_type, "工具扫描")
        self.assertEqual(csv_alert.attack_ip, "9.9.9.9")
        self.assertEqual(csv_alert.target_ip, "10.4.5.6")
        self.assertEqual(csv_alert.event_type, "暴力破解")

    def test_local_xlsx_key_value_and_table_files_are_supported(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            key_value = root / "key_value.xlsx"
            workbook = Workbook()
            sheet = workbook.active
            sheet.append(["攻击IP", "203.0.113.9"])
            sheet.append(["目的IP", "10.2.9.8"])
            sheet.append(["攻击名称", "Nmap扫描"])
            workbook.save(key_value)

            table = root / "table.xlsx"
            workbook = Workbook()
            sheet = workbook.active
            sheet.append(["src_ip", "dst_ip", "attack_name", "result"])
            sheet.append(["198.51.100.4", "10.4.5.6", "SSH暴力破解", "企图"])
            workbook.save(table)

            first = parse_local_file(key_value)
            second = parse_local_file(table)
            evidence = file_to_text(table)
        self.assertEqual(first.attack_ip, "203.0.113.9")
        self.assertEqual(first.target_ip, "10.2.9.8")
        self.assertEqual(second.attack_ip, "198.51.100.4")
        self.assertEqual(second.target_ip, "10.4.5.6")
        self.assertEqual(second.attack_result, "失败")
        self.assertIn("src_ip: 198.51.100.4", evidence)

    def test_local_inline_key_values_infer_attack_without_ai(self) -> None:
        text = (
            "timestamp=2026-08-10T12:34:56Z src_ip=203.0.113.8 "
            "dst_ip=10.2.8.9 severity=medium "
            "request_uri=/download?file=../../etc/passwd result=blocked"
        )
        alert = smart_extract(
            settings={"analysis_mode": "local", "ai_enabled": True}, text=text
        )
        self.assertEqual(alert.time, "2026-08-10 12:34:56")
        self.assertEqual(alert.attack_ip, "203.0.113.8")
        self.assertEqual(alert.target_ip, "10.2.8.9")
        self.assertEqual(alert.alert_level, "中危")
        self.assertEqual(alert.attack_name, "目录遍历攻击(通用)")
        self.assertEqual(alert.event_type, "目录遍历")
        self.assertEqual(alert.attack_result, "失败")

    def test_local_html_embedded_json_is_available_to_rule_parser(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "dashboard.html"
            path.write_text(
                """<html><body><script type="application/json">
                {"event":{"src_ip":"198.51.100.9","dst_ip":"10.3.4.5",
                "rule_name":"Nmap scan","severity":"low"}}
                </script></body></html>""",
                encoding="utf-8",
            )
            alert = local_extract_text_or_html(path=path)
        self.assertEqual(alert.attack_ip, "198.51.100.9")
        self.assertEqual(alert.target_ip, "10.3.4.5")
        self.assertEqual(alert.event_type, "工具扫描")
        self.assertEqual(alert.alert_level, "低危")

    def test_timestamp_is_not_used_as_a_rule_name(self) -> None:
        alert = parse_text(
            "2026-08-10 12:34:56 src_ip=1.2.3.4 dst_ip=10.1.2.3 status=blocked"
        )
        self.assertNotIn("08-10", alert.attack_name)

    def test_order_build_returns_copyable_standard_text(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            engine = self._empty_engine(root)
            order = assemble_order(
                {"number": "0807-165", "attack_ip": "8.8.8.8", "target_ip": "10.2.3.4", "attack_name": "SSH暴力破解攻击"},
                engine,
            )
            self.assertIn("封禁 8.8.8.8", order.advice)
            text = order.to_markdown()
            self.assertIn("编号：0807-165", text)
            self.assertFalse((root / "orders.md").exists())

    def test_advice_has_only_public_block_or_internal_authorization_forms(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            engine = self._empty_engine(Path(tmp))
            public = assemble_order({"attack_ip": "8.8.8.8", "attack_result": "企图"}, engine)
            internal = assemble_order({"attack_ip": "10.2.3.4", "attack_result": "企图"}, engine)
            self.assertEqual(public.attack_result, "失败")
            self.assertEqual(public.advice, "封禁 8.8.8.8")
            self.assertEqual(
                internal.advice,
                "核实 10.2.3.4 的任务授权情况；已授权则加白，未授权则隔离并排查源主机",
            )
            ioc = assemble_order({"attack_ip": "203.0.113.8", "domain_url": "zdg16881988.com"}, engine)
            self.assertIn("封禁 203.0.113.8", ioc.advice)
            self.assertIn("封禁 zdg16881988.com", ioc.advice)
            ai_advice = assemble_order(
                {"attack_ip": "203.0.113.8", "domain_url": "zdg16881988.com", "advice": "封禁 203.0.113.8"},
                engine, auto_advice=False,
            )
            self.assertIn("封禁 zdg16881988.com", ai_advice.advice)

    def test_batch_second_run_does_not_duplicate_generated_output(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            engine = self._empty_engine(root)
            history = HistoryStore(root / "history.json")
            jobs = jobs_from_text_blob(self.SAMPLE)
            settings = {"ai_enabled": False, "number_seq": {}, "active_template": "默认工单字段"}
            output = root / "orders.md"
            first = process_batch(
                jobs, wl=engine, history=history, settings=settings,
                source="自定义监测平台", event_level="五级", date_mmdd="0807",
            )
            second = process_batch(
                jobs, wl=engine, history=history, settings=settings,
                source="自定义监测平台", event_level="五级", date_mmdd="0807",
            )
            self.assertEqual(first.written, 1)
            self.assertEqual(second.written, 0)
            self.assertFalse(output.exists())
            self.assertEqual(jobs[0].order_md.count("编号："), 1)

    def test_job_path_deduplication(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "a.txt"
            path.write_text(self.SAMPLE, encoding="utf-8")
            self.assertEqual(len(jobs_from_paths([str(path), str(path)])), 1)

    def test_history_xlsx_reload_and_match(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            workbook = Workbook()
            sheet = workbook.active
            sheet.append(["序号", "编号", "监测来源", "时间", "攻击IP", "目标IP", "XFF", "域名URL", "攻击名称", "事件类型", "上报人员"])
            sheet.append([1, "0807-001", "自定义监测平台", "2026-08-07 20:00:00", "8.8.8.8", "10.2.3.4", "", "", "SSH暴力破解攻击", "暴力破解", "A"])
            xlsx = root / "history.xlsx"
            workbook.save(xlsx)
            store = HistoryStore(root / "history.json")
            self.assertEqual(store.reload_from_xlsx(xlsx), 1)
            hits = store.find_duplicates(["8.8.8.8"], target_ip="10.2.3.4", attack_name="SSH暴力破解攻击")
            self.assertEqual(hits[0].code, "0807-001")


if __name__ == "__main__":
    unittest.main()
