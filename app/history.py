"""历史告警去重：已经上报或处置的同类告警不再重复产出。"""

from __future__ import annotations

import json
import re
from io import BytesIO
from dataclasses import asdict, dataclass
from datetime import datetime
from pathlib import Path
from typing import Any

from .constants import HISTORY_CACHE_PATH
from .io_utils import atomic_write_text
from .default_whitelist import active_project_profile_path
from .whitelist import extract_ips

try:
    import openpyxl
except ImportError:  # pragma: no cover
    openpyxl = None


@dataclass
class HistoryHit:
    sheet: str
    code: str
    time: str
    attack_ip: str
    target_ip: str
    attack_name: str
    reporter: str
    reason: str


def _clean(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _feature(value: str) -> str:
    return re.sub(r"[^0-9a-z\u4e00-\u9fff]+", "", (value or "").lower())


def _similar(left: str, right: str, minimum: int = 4) -> bool:
    a, b = _feature(left), _feature(right)
    return bool(a and b and min(len(a), len(b)) >= minimum and (a in b or b in a))


class HistoryStore:
    def __init__(self, cache_path: Path | None = None) -> None:
        self.cache_path = cache_path or HISTORY_CACHE_PATH
        self.records: list[dict[str, Any]] = []
        self._ip_index: dict[str, list[dict[str, Any]]] = {}
        self.load_cache()

    def load_cache(self) -> None:
        project_path = None if self.cache_path != HISTORY_CACHE_PATH else active_project_profile_path()
        if project_path:
            try:
                project = json.loads(project_path.read_text(encoding="utf-8"))
                raw = project.get("history_cache") if isinstance(project, dict) else None
                self.records = raw if isinstance(raw, list) else []
            except Exception:
                self.records = []
        elif self.cache_path.exists():
            try:
                raw = json.loads(self.cache_path.read_text(encoding="utf-8"))
                self.records = raw if isinstance(raw, list) else []
            except Exception:
                self.records = []
        self._rebuild_index()

    def save_cache(self) -> None:
        project_path = None if self.cache_path != HISTORY_CACHE_PATH else active_project_profile_path()
        if project_path:
            project = json.loads(project_path.read_text(encoding="utf-8"))
            project["history_cache"] = self.records
            atomic_write_text(project_path, json.dumps(project, ensure_ascii=False, indent=2) + "\n")
        else:
            atomic_write_text(self.cache_path, json.dumps(self.records, ensure_ascii=False, indent=2) + "\n")

    def _rebuild_index(self) -> None:
        index: dict[str, list[dict[str, Any]]] = {}
        for record in self.records:
            for ip in extract_ips(f"{record.get('attack_ip', '')} {record.get('xff', '')}"):
                index.setdefault(ip, []).append(record)
        self._ip_index = index

    def reload_from_xlsx(self, xlsx_path: str | Path) -> int:
        if openpyxl is None:
            raise RuntimeError("缺少 openpyxl，无法读取历史表")
        path = Path(xlsx_path)
        if not path.exists():
            raise FileNotFoundError(f"历史表不存在: {path}")

        workbook = openpyxl.load_workbook(path, data_only=True, read_only=True)
        try:
            records = self._records_from_workbook(workbook)
        finally:
            workbook.close()

        self.records = records
        self._rebuild_index()
        self.save_cache()
        return len(records)

    def _records_from_workbook(self, workbook: Any, sheet_label: str = "") -> list[dict[str, Any]]:
        records: list[dict[str, Any]] = []
        for sheet_name in workbook.sheetnames:
            if str(sheet_name).startswith("Wps"):
                continue
            sheet = workbook[sheet_name]
            rows = sheet.iter_rows(values_only=True)
            headers: list[str] | None = None
            header_indexes: dict[str, list[int]] = {}
            for row in rows:
                values = list(row)
                if headers is None:
                    labels = [_clean(v) for v in values]
                    if "序号" in labels and "攻击IP" in labels and "编号" in labels:
                        headers = labels
                        for index, label in enumerate(labels):
                            if label:
                                header_indexes.setdefault(label, []).append(index)
                    continue

                def value(name: str, occurrence: int = 0) -> str:
                    indexes = header_indexes.get(name) or []
                    if occurrence >= len(indexes):
                        return ""
                    idx = indexes[occurrence]
                    return _clean(values[idx] if idx < len(values) else None)

                code = value("编号")
                attack_ip = value("攻击IP")
                if not code or not extract_ips(attack_ip):
                    continue
                status_values = [
                    value("上报人员"), value("上报时间"), value("处置建议修订结果"),
                    value("是否转入事件跟踪"), value("攻击IP是否封禁"), value("研判人员"),
                    value("情报人员"), value("备注"), value("备注", 1),
                ]
                if not any(status_values):
                    continue
                all_fields: dict[str, str] = {}
                duplicates: dict[str, int] = {}
                for index, header in enumerate(headers):
                    if not header:
                        continue
                    duplicates[header] = duplicates.get(header, 0) + 1
                    field_name = header if duplicates[header] == 1 else f"{header}({duplicates[header]})"
                    raw_value = values[index] if index < len(values) else None
                    all_fields[field_name] = self._fmt_time(raw_value)
                records.append({
                    "sheet": sheet_label or _clean(sheet_name),
                    "code": code,
                    "source": value("监测来源"),
                    "time": self._fmt_time(values[(header_indexes.get("时间") or [0])[0]]),
                    "attack_ip": attack_ip,
                    "target_ip": value("目标IP"),
                    "xff": value("XFF"),
                    "domain_url": value("域名URL"),
                    "attack_name": value("攻击名称"),
                    "event_type": value("事件类型"),
                    "reporter": value("上报人员") or value("研判人员"),
                    "all_fields": all_fields,
                })
        return records

    @staticmethod
    def _record_key(record: dict[str, Any]) -> tuple[str, ...]:
        code = _clean(record.get("code"))
        if code:
            return ("code", code.casefold())
        return (
            "event", _clean(record.get("time")), _clean(record.get("attack_ip")),
            _clean(record.get("target_ip")), _feature(_clean(record.get("attack_name"))),
        )

    def merge_records(self, incoming: list[dict[str, Any]]) -> tuple[int, int, int]:
        """Append unseen records and refresh matching records from a source workbook."""
        index = {self._record_key(record): pos for pos, record in enumerate(self.records)}
        added = updated = 0
        for raw in incoming:
            record = {
                key: value if isinstance(value, (dict, list)) else _clean(value)
                for key, value in raw.items()
            }
            key = self._record_key(record)
            previous = index.get(key)
            if previous is None:
                self.records.append(record)
                index[key] = len(self.records) - 1
                added += 1
                continue
            old = self.records[previous]
            merged = dict(old)
            merged.update({key: value for key, value in record.items() if value})
            if merged != old:
                self.records[previous] = merged
                updated += 1
        if added or updated:
            self._rebuild_index()
            self.save_cache()
        return added, updated, len(self.records)

    def merge_from_xlsx_bytes(self, payload: bytes, sheet_label: str = "") -> tuple[int, int, int]:
        if openpyxl is None:
            raise RuntimeError("缺少 openpyxl，无法读取历史表")
        try:
            workbook = openpyxl.load_workbook(BytesIO(payload), data_only=True, read_only=True)
        except Exception as exc:
            raise ValueError("下载的 Excel 无法读取") from exc
        try:
            records = self._records_from_workbook(workbook, sheet_label=sheet_label)
        finally:
            workbook.close()
        if not records:
            raise ValueError("Excel 中未找到有效的告警跟踪记录")
        return self.merge_records(records)

    def confirm_order(self, order: Any, *, raw_result: str = "") -> tuple[bool, int]:
        """Persist a confirmed generated work order in the local tracking cache."""
        full = order.to_dict() if hasattr(order, "to_dict") else {}
        record = {
            "sheet": "工具确认",
            "code": _clean(getattr(order, "number", "")),
            "source": _clean(getattr(order, "source", "")),
            "time": _clean(getattr(order, "time", "")),
            "attack_ip": _clean(getattr(order, "attack_ip", "")),
            "target_ip": _clean(getattr(order, "target_ip", "")),
            "xff": _clean(getattr(order, "xff", "")),
            "domain_url": _clean(getattr(order, "domain_url", "")),
            "attack_name": _clean(getattr(order, "attack_name", "")),
            "event_type": _clean(getattr(order, "event_type", "")),
            "reporter": "工具确认",
            "raw_result": _clean(raw_result),
            "all_fields": full if isinstance(full, dict) else {},
        }
        if not record["code"]:
            raise ValueError("工单编号为空，无法确认")
        before = len(self.records)
        added, updated, total = self.merge_records([record])
        return bool(added or updated or len(self.records) != before), total

    def search(self, keyword: str) -> list[dict[str, Any]]:
        query = _clean(keyword).casefold()
        if not query:
            return list(self.records)
        return [
            record for record in self.records
            if query in json.dumps(record, ensure_ascii=False).casefold()
        ]

    def find_exact_code(self, code: str) -> dict[str, Any] | None:
        wanted = _clean(code).casefold()
        if not wanted:
            return None
        return next(
            (record for record in self.records if _clean(record.get("code")).casefold() == wanted),
            None,
        )

    @staticmethod
    def _fmt_time(value: Any) -> str:
        if value is None:
            return ""
        if isinstance(value, datetime):
            return value.strftime("%Y-%m-%d %H:%M:%S")
        return str(value).strip()

    def find_duplicates(
        self,
        attack_ips: list[str],
        target_ip: str = "",
        attack_name: str = "",
        xff: str = "",
        domain_url: str = "",
        event_type: str = "",
    ) -> list[HistoryHit]:
        """以攻击方 IP 为主键，并要求目标、URL 或攻击特征至少一项相符。"""
        candidate_ips = []
        for ip in [*attack_ips, *extract_ips(xff)]:
            if ip not in candidate_ips:
                candidate_ips.append(ip)
        current_targets = set(extract_ips(target_ip))
        current_url = (domain_url or "").strip().lower().rstrip("/")

        hits: list[HistoryHit] = []
        seen: set[tuple[str, str, str]] = set()
        for ip in candidate_ips:
            for record in self._ip_index.get(ip, []):
                key = (record.get("sheet", ""), record.get("code", ""), record.get("time", ""))
                if key in seen:
                    continue
                rec_targets = set(extract_ips(record.get("target_ip", "")))
                rec_url = (record.get("domain_url", "") or "").strip().lower().rstrip("/")
                target_match = bool(current_targets and rec_targets and current_targets & rec_targets)
                url_match = bool(
                    current_url and rec_url and min(len(current_url), len(rec_url)) >= 8
                    and (current_url in rec_url or rec_url in current_url)
                )
                name_match = _similar(attack_name, record.get("attack_name", ""))
                type_match = _similar(event_type, record.get("event_type", ""), minimum=3)
                if not (target_match or url_match or name_match or type_match):
                    continue

                matched_features = []
                if target_match:
                    matched_features.append("目标IP一致")
                if url_match:
                    matched_features.append("URL一致")
                if name_match:
                    matched_features.append("攻击名称相似")
                if type_match:
                    matched_features.append("事件类型相似")
                seen.add(key)
                hits.append(HistoryHit(
                    sheet=record.get("sheet", ""),
                    code=record.get("code", ""),
                    time=record.get("time", ""),
                    attack_ip=record.get("attack_ip", ""),
                    target_ip=record.get("target_ip", ""),
                    attack_name=record.get("attack_name", ""),
                    reporter=record.get("reporter", ""),
                    reason=f"攻击方IP {ip} 已处置；" + "、".join(matched_features),
                ))
        return hits

    def find_related(
        self,
        attack_ip: str = "",
        target_ip: str = "",
        attack_name: str = "",
        event_type: str = "",
    ) -> list[HistoryHit]:
        """Find prior submissions involving either side of the current event.

        This intentionally scans the cached rows instead of only the attack-IP
        index: a new attacker against an already tracked target is useful
        context for the online re-submission judgment.
        """
        attack_ips = set(extract_ips(attack_ip))
        target_ips = set(extract_ips(target_ip))
        hits: list[HistoryHit] = []
        for record in self.records:
            old_attack = set(extract_ips(record.get("attack_ip", "")))
            old_target = set(extract_ips(record.get("target_ip", "")))
            same_attack = bool(attack_ips & old_attack)
            same_target = bool(target_ips & old_target)
            name_match = _similar(attack_name, record.get("attack_name", ""))
            type_match = _similar(event_type, record.get("event_type", ""), minimum=3)
            if not (same_attack or same_target or name_match or type_match):
                continue
            reasons = []
            if same_attack:
                reasons.append("攻击IP相同")
            if same_target:
                reasons.append("目标IP相同")
            if name_match:
                reasons.append("攻击名称相似")
            if type_match:
                reasons.append("事件类型相似")
            hits.append(HistoryHit(
                sheet=record.get("sheet", ""), code=record.get("code", ""),
                time=record.get("time", ""), attack_ip=record.get("attack_ip", ""),
                target_ip=record.get("target_ip", ""), attack_name=record.get("attack_name", ""),
                reporter=record.get("reporter", ""), reason="、".join(reasons),
            ))
        return hits

    @staticmethod
    def as_dicts(hits: list[HistoryHit]) -> list[dict[str, Any]]:
        return [asdict(hit) for hit in hits]
