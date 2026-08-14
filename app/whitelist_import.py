"""从公司出口信息表同步白名单规则。"""

from __future__ import annotations

import re
import csv
import json
import io
from pathlib import Path
from typing import Any

try:
    import openpyxl
except ImportError:  # pragma: no cover
    openpyxl = None


_IP_TOKEN_RE = re.compile(
    r"(?<![\d.])(?:\d{1,3}\.){3}\d{1,3}(?:/\d{1,2})?(?![\d.])"
    r"|(?<![\d.])(?:\d{1,3}\.){2}\d{1,3}\.[xX](?![\w.])"
)


def _normalize_rule(token: str) -> str:
    token = token.strip()
    if token.lower().endswith(".x"):
        return token[:-1] + "0/24"
    return token


def extract_rules_from_text(text: str, *, source: str = "导入文件", reason: str = "文件导入") -> list[dict[str, str]]:
    """Extract IP/CIDR rules from arbitrary text and merge duplicate tokens."""
    found: dict[str, dict[str, str]] = {}
    for match in _IP_TOKEN_RE.finditer(text or ""):
        token = _normalize_rule(match.group(0))
        before = (text[max(0, match.start() - 80):match.start()] or "").strip()
        line = before.rsplit("\n", 1)[-1].strip(" ：:,-")
        item = {"rule": token, "reason": line[-40:] or reason, "source": source}
        found.setdefault(token, item)
    return list(found.values())


def extract_rules_from_file(path: str | Path) -> list[dict[str, str]]:
    """Read TXT/CSV/JSON/HTML/XLSX whitelist files without replacing existing rules."""
    path = Path(path)
    if not path.exists() or not path.is_file():
        raise FileNotFoundError(f"白名单文件不存在: {path}")
    suffix = path.suffix.casefold()
    if suffix in {".xlsx", ".xlsm"}:
        # Workbooks may also contain company department allocations. Never
        # scan every sheet as a fallback or those attribution-only networks
        # could become exemption whitelist entries.
        return extract_network_rules(path)
    raw = path.read_bytes()
    for encoding in ("utf-8-sig", "gb18030", "utf-16"):
        try:
            text = raw.decode(encoding)
            break
        except UnicodeDecodeError:
            text = raw.decode("utf-8", errors="ignore")
    if suffix == ".json":
        try:
            text = json.dumps(json.loads(text), ensure_ascii=False)
        except Exception:
            pass
    elif suffix in {".csv", ".tsv"}:
        try:
            dialect = csv.excel_tab if suffix == ".tsv" else csv.Sniffer().sniff(text[:4096])
            rows = csv.reader(io.StringIO(text), dialect=dialect)
            text = "\n".join(", ".join(row) for row in rows)
        except Exception:
            pass
    elif suffix in {".html", ".htm"}:
        try:
            from bs4 import BeautifulSoup
            text = BeautifulSoup(text, "lxml").get_text("\n", strip=True)
        except Exception:
            text = re.sub(r"<[^>]+>", "\n", text)
    return extract_rules_from_text(text, source="白名单文件")


def merge_rules_from_file(engine: Any, path: str | Path) -> int:
    """Incrementally import, normalize, and deduplicate rules in an engine."""
    imported = extract_rules_from_file(path)
    if not imported:
        return 0
    return engine.merge_rules(imported)


def _reason_for(header: str, section: str, cell_text: str, token: str) -> str:
    suffix = cell_text.replace(token, "").strip(" ：:（）()")
    if suffix and not re.search(r"[\d/]", suffix):
        return suffix
    if "出口" in header and section:
        return f"{header.replace('出口', '')}{section}出口"
    return section or header or "公司白名单"


def extract_network_rules(xlsx_path: str | Path) -> list[dict[str, str]]:
    """只读取明确的出口/DNS/VPN/设备白名单页。"""
    if openpyxl is None:
        raise RuntimeError("缺少 openpyxl，无法读取白名单源表")
    path = Path(xlsx_path)
    if not path.exists():
        raise FileNotFoundError(f"白名单源表不存在: {path}")

    workbook = openpyxl.load_workbook(path, data_only=True, read_only=True)
    try:
        sheet = None
        for candidate in workbook.worksheets:
            first = [str(v or "").strip() for v in next(candidate.iter_rows(max_row=1, values_only=True))]
            if any(v in {"DNS", "VPN", "防火墙", "扫描器", "渗透固定ip"} for v in first):
                sheet = candidate
                break
        if sheet is None:
            raise ValueError("未找到包含出口/DNS/VPN/设备信息的工作表")

        rows = list(sheet.iter_rows(values_only=True))
        headers = [str(v or "").strip() for v in rows[0]]
        found: dict[str, dict[str, str]] = {}
        reason_priority = {"扫描器": 7, "渗透固定ip": 6, "防火墙": 5, "DNS": 4, "VPN": 3}

        for col, header in enumerate(headers):
            section = ""
            for row in rows[1:]:
                value: Any = row[col] if col < len(row) else None
                text = str(value or "").strip()
                if not text:
                    continue
                tokens = [_normalize_rule(m.group(0)) for m in _IP_TOKEN_RE.finditer(text)]
                if not tokens:
                    if len(text) <= 24:
                        section = text.rstrip("：:")
                    continue
                for token in tokens:
                    reason = _reason_for(header, section, text, token)
                    item = {"rule": token, "reason": reason, "source": "白名单表"}
                    previous = found.get(token)
                    if previous is None or reason_priority.get(reason, 1) > reason_priority.get(previous["reason"], 1):
                        found[token] = item
        return list(found.values())
    finally:
        workbook.close()


def sync_network_rules(engine: Any, xlsx_path: str | Path) -> int:
    """Replace imported whitelist-table rules and preserve other entries."""
    imported = extract_network_rules(xlsx_path)
    data = {"version": 1, "description": "白名单IP/网段", "rules": [], "manual": []}
    if engine.path.exists():
        import json

        data.update(json.loads(engine.path.read_text(encoding="utf-8")))
    preserved = [
        r for r in (data.get("rules") or [])
        if r.get("source") not in {"公司网段表", "白名单表"}
    ]
    engine.save(rules=preserved + imported, manual=list(data.get("manual") or []))
    return len(imported)
